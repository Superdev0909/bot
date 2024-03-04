from selenium import webdriver
from selenium.common.exceptions import *
import json
import func
import os
from time import sleep
import openpyxl
import pandas as pd
from datetime import datetime

def submit():
    with open('verified_emails.json', 'r') as file:
        accounts = json.load(file)
    for i in range(count):
        for name in profiles:
            with open(f'profiles/{name}.json', 'r') as fp:
                profile = json.load(fp)
            NotAppliedExist = False
            for index_account in range(len(accounts)):
                if accounts[index_account]['applied'] == False:
                    account_index = index_account
                    email = accounts[account_index]['email']
                    # password = accounts[account_index]['password']
                    password = "superstar_0909"
                    NotAppliedExist = True
                    break
            if NotAppliedExist:
                driver = webdriver.Chrome()
                driver.maximize_window()
                action = webdriver.ActionChains(driver)
                action = webdriver.common.action_chains.ActionChains(driver)

                func.login(driver, email, password)
                func.get_started(driver)
                
                accounts[account_index]['applied'] = True
                with open('verified_emails.json', 'w') as file:
                    json.dump(accounts, file)

                func.select_experience(driver)
                func.select_what_is_my_goal(driver)
                func.select_work_preference(driver)
                func.select_manualmode(driver)
                func.add_professional(driver, profile['professional'])
                func.add_experience(driver, profile['work_experience'])
                func.add_education(driver, profile['education'], action)
                func.add_language(driver, profile['languages'])
                func.add_skills(driver, profile['skills'])
                func.add_overview(driver, profile['overview'])
                func.add_service(driver, profile['services'])
                func.add_rate(driver, profile['hour_rate'])
                func.add_photo_others(driver, profile['photo_others'], action)
                # func.submit_profile(driver)
                driver.quit()
            else:
                print('All accounts was applied.')

    with open('verified_emails.json') as file:
        data = json.load(file)

    # Create a list to store the values of the specified key
    email = []

    # Iterate over each item in the JSON data
    for item in data:
        # Get the value of the specified key
        email.append(item['email'])

    # Create a DataFrame from the values
    df = pd.DataFrame(email)
    df.columns = [None]  # Set the column name to None

    # Save the DataFrame to an Excel file
    current_datetime = datetime.now().strftime("%m%d%H%M")
    filename = os.path.join(".", current_datetime + ".xlsx")

    df.to_excel(filename, index=False, header=False)
    print("Exported successfully to:", filename)

if __name__ == "__main__":
    response_1 = input('Would you like to create new emails? (y/n) :')
    response_2 = input('Would you like to apply your profile to new emails? (y/n) :')
    count = int(input('Enter number of accounts for each profile. '))
    if count > 0:
        subfix = ''
        while True:
            text = input(f'Enter profiles to be used {subfix}.')
            profiles = text.split(' ')
            exists = True
            for name in profiles:
                if not os.path.exists(f'profiles/{name}.json'):
                    exists = False
            if exists:
                break
            else:
                subfix = 'again'
            sleep(1)
        profiles = text.split(' ')
        if response_1 == 'y':
            excel = openpyxl.load_workbook("email.xlsx")
            sheet = excel.active
            excel_file = "email.xlsx"
            rowCount = sheet.max_row + 1
            emails = []
            for row in range(1, rowCount):
                email = sheet.cell(row=row, column=1).value
                emails.append({"name" : email, "verified" : False})
            print(emails)
            # emails = func.get_email(count)
            func.verify_email(emails, profiles, count)
        if response_2 == 'y':
            submit()
    print('done')